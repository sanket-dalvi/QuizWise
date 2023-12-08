from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from QuizWise.auth_decorator import examiner_required
from django.contrib import messages
from .models import Category, QuestionType, QuestionOption, Question, CategoryQuestionMap, Quiz, QuizQuestion, Group, GroupExamineeMapping
from QuizParticipant.models import UserQuizScore, UserQuizStatus
from QuizWise.models import User
from django.db.models import Q
from django.core.exceptions import ValidationError
import json
from django.db.models import Prefetch, Sum, F, FloatField
from django.http import JsonResponse
import statistics
import matplotlib.pyplot as plt
import base64
from io import BytesIO
import seaborn as sns
import pandas as pd
import random
from .forms import FileUploadForm
import os, io
from django.conf import settings
from QuizWise.notification import notification
from openpyxl import load_workbook
from QuizWise.decorators import log_view


@log_view
@examiner_required
def home(request):
    # Get quiz metrics for the current user and render the home template
    context = get_quiz_metrics(request)
    context['first_name'] = request.user.first_name
    context['last_name'] = request.user.last_name
    context['email'] = request.user.email
    context['contact'] = request.user.contact

    return render(request, "QuizCreator/home.html", context)

def get_quiz_metrics(request, quiz=None):
    current_user = request.user
    recent_quiz_score = UserQuizScore.objects.filter(
        quiz__created_by=current_user
    ).order_by('-timestamp').first()

    if recent_quiz_score:
        recent_quiz = quiz
        if not quiz:
            recent_quiz = recent_quiz_score.quiz
        user_total_scores = UserQuizScore.objects.filter(quiz=recent_quiz).values('user').annotate(
            total_score=Sum('score')
        )
        total_scores = [user['total_score'] for user in user_total_scores]

        mean_score = statistics.mean(total_scores)
        median_score = statistics.median(total_scores)
        mode_score = statistics.mode(total_scores)
        min_score = min(user_total_scores, key=lambda x: x['total_score'])['total_score']
        max_score = max(user_total_scores, key=lambda x: x['total_score'])['total_score']

        recent_quiz_metrics = {
            'quiz_name': recent_quiz.name,
            'duration': recent_quiz.duration,
            'total_questions': recent_quiz.total_questions,
            'min': min_score,
            'max': max_score,
            'mean': mean_score,
            'median': median_score,
            'mode': mode_score
        }

    else:
        recent_quiz_metrics = None

    context = {
        'recent_quiz_metrics': recent_quiz_metrics,
    }
    return context

@log_view
@examiner_required
def profile(request):
    # Display and update user profile information.
    if request.method == "POST":
        try:
            # Extract and update user details from the POST request
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            contact = request.POST.get('contact')
            email_notification = request.POST.get('email_notification')
            mobile_notification = request.POST.get('mobile_notification')
            email_notification = email_notification == 'on'  
            mobile_notification = mobile_notification == 'on'
            user = get_object_or_404(User, pk=request.user.id)
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.contact = contact
            user.email_notification = email_notification
            user.mobile_notification = mobile_notification
            user.save()
            # Display success message upon successful update
            messages.success(request, "User Details Updated Successfully")
        except Exception as e:
            messages.error(request, f"ERROR : {str(e)}")

    # Retrieve user details for rendering the profile template
    user = get_object_or_404(User, pk=request.user.id)
    context = {
        "first_name" : user.first_name,
        "last_name" : user.last_name,
        "email" : user.email,
        "contact" : user.contact,
        "email_notification" : user.email_notification,
        "mobile_notification" : user.mobile_notification
    }
    return render(request, "QuizCreator/profile.html", context)


@log_view
@examiner_required
def scores(request):
    # Display and calculate scores for quiz submissions.
    quizzes = Quiz.objects.filter(created_by=request.user)

    if request.method == 'POST':
        # Handle POST request to display scores for a selected quiz
        selected_quiz_id = request.POST.get('select-quiz')
        selected_quiz = Quiz.objects.filter(id=selected_quiz_id).first()       
        user_quiz_scores = UserQuizScore.objects.filter(quiz=selected_quiz)        
        distinct_users = user_quiz_scores.values_list('user', flat=True).distinct()
        quiz_data = []

        # Calculate and organize user scores and related information
        for user_id in distinct_users:
            user_scores = user_quiz_scores.filter(user_id=user_id)
            user_total_score = user_scores.aggregate(total_score=Sum('score'))['total_score'] or 0
            total_questions = QuizQuestion.objects.filter(quiz=selected_quiz).count()
            percentage = (
                user_total_score * 100.0 / (selected_quiz.total_questions * 1.0)
            ) if total_questions > 0 else 0
            user_data = {
                'username': user_scores.first().user.username,
                'email': user_scores.first().user.email,
                'total_questions': total_questions,
                'score': user_total_score,
                'percentage': percentage
            }
            quiz_data.append(user_data)
        quiz_data.sort(key=lambda data: data['score'], reverse=True)

        # Prepare context for rendering the template
        context = {
            'total_submissions' : len(quiz_data),
            'quiz_data': quiz_data,
            'selected_quiz': selected_quiz,
            'quizzes' : quizzes,
        }

        # Include additional metrics context
        if len(quiz_data) > 0:
            metric_context = get_quiz_metrics(request, selected_quiz)
            context.update(metric_context)
        return render(request, "QuizCreator/scores.html", context)
    
    # Default context for rendering the template without quiz data
    context = {
        'quiz_data': [],
        'selected_quiz': None,
        'quizzes' : quizzes,
    }
    return render(request, "QuizCreator/scores.html", context)


@log_view
@examiner_required
def quiz_history(request):
    return render(request, "QuizCreator/quiz_history.html")

@log_view
@examiner_required
def create_question(request):
    # Handle the creation of a new question.
    if request.method == "POST":
        # Retrieve question details from the form submission
        question_text = request.POST.get('question')
        question_type_code = request.POST.get('type')        
        radio_options = request.POST.getlist('radio-group')
        checkbox_options = request.POST.getlist('checkbox-group')
        selected_radio = request.POST.get('radio-group')
        selected_checkbox = request.POST.getlist('checkbox-group')
        answer = request.POST.get("free-text-answer")
        options_json = request.POST.get('options')
        options = json.loads(options_json) if options_json else []

        # Validate input for question text and question type
        if not question_text or not question_type_code:
            messages.error(request, "Please fill in all fields.")
            return redirect('create_question')
        
        # Validate correct answers for radio buttons and checkboxes
        if question_type_code in ['RB', 'CB']:
            if not radio_options and question_type_code == 'RB':
                messages.error(request, "Please select correct answer option for the radio button.")
                return redirect('create_question')           
            if not checkbox_options and question_type_code == 'CB':
                messages.error(request, "Please select correct answer options for the checkboxes.")
                return redirect('create_question')
        
        # Validate selected answer for radio buttons and checkboxes
        if question_type_code in ['RB', 'CB']:
            if not selected_radio and question_type_code == 'RB':
                messages.error(request, "Please select a correct answer for the radio button.")
                return redirect('create_question')           
            if not options:
                messages.error(request, "Please add options.")
                return redirect('create_question')
            if question_type_code == 'FT':
                options = []
            elif question_type_code == 'RB':
                answer = selected_radio if selected_radio else ''
            elif question_type_code == 'CB':
                answer = selected_checkbox if selected_checkbox else []

        current_user = request.user
        try:  
            # Create a new Question object       
            question = Question.objects.create(
                question=question_text,
                type=QuestionType.objects.get(type_code=question_type_code),
                answer=answer,  
                created_by=current_user,
            )

            # Create QuestionOption objects for multiple-choice questions          
            for option in options:
                QuestionOption.objects.create(question=question, option=option)          
            messages.success(request, "Question created successfully.")
            return redirect('create_question')
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('create_question')
   
    # Render the 'create_question.html' template with relevant context for question creation
    current_user = request.user
    question_types = QuestionType.objects.all()
    categories = Category.objects.filter(
        Q(created_by=current_user) | Q(visible_to_others=True)
    )
    return render(request, "QuizCreator/create_question.html", {'question_types': question_types})

@log_view
@examiner_required
def question_file_download(request):
    # Handle file download for the question upload template.
    file_path = os.path.join(settings.MEDIA_ROOT, 'QUESTION_UPLOAD_TEMPLATE.xlsx')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/force-download')
            response['Content-Disposition'] = 'attachment; filename=QUESTION_UPLOAD_TEMPLATE.xlsx'
            return response
    return HttpResponse("File not found")

@log_view
@examiner_required
def question_file_upload(request):
    # Handle file upload for creating questions from an Excel file. 
    if request.method == 'POST':
        # Validate and process the uploaded file using the FileUploadForm
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Retrieve the uploaded file and save it to the 'upload' directory
            uploaded_file = request.FILES['file']
            file_path = os.path.join(settings.MEDIA_ROOT, 'upload', uploaded_file.name)
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            # Load the workbook and retrieve the active sheet
            wb = load_workbook(file_path)
            sheet = wb.active

            # Lists to store created questions and question options
            questions = []
            question_options = []

            # Iterate through rows in the sheet and process each question
            for row in sheet.iter_rows(min_row=11, values_only=True):
                if len(row) >= 3:
                    question_text = row[0].strip() if isinstance(row[0], str) else row[0]
                    question_type_name = row[1].strip() if isinstance(row[1], str) else row[1]
                    correct_answer = row[2].strip() if isinstance(row[2], str) else row[2]
                    options = row[3] if len(row) >= 4 else ''  
                    question_type = QuestionType.objects.filter(type_name=question_type_name).first()
                    if not question_type:
                        continue  # Skip if the question type doesn't exist

                    # Validation: Check if the correct answer is among the options for MCQ type
                    if question_type.type_code != 'FT' and correct_answer.lower() not in [option.strip().lower() for option in options.split(',')]:
                        continue  # Skip if correct answer not in options for MCQ type

                    # Create a new Question object
                    new_question = Question(
                        question=question_text,
                        type=question_type,
                        answer=correct_answer,
                        created_by=request.user 
                    )

                    # Process and create QuestionOption objects for multiple-choice questions
                    if question_type.type_code != 'FT':
                        if options:
                            options = options.split(",")
                            for option in options:
                                option = option.strip()
                                question_option = QuestionOption(
                                    question = new_question,
                                    option = option
                                )
                                question_options.append(question_option)
                    questions.append(new_question)
            Question.objects.bulk_create(questions)
            QuestionOption.objects.bulk_create(question_options)

            # Display success message, remove the uploaded file, and redirect to 'create_question' page
            messages.success(request, "Questions Uploaded Successfully")
            os.remove(file_path)
            return redirect('create_question')
    else:
        form = FileUploadForm()
    return render(request, 'QuizCreator/create_question.html', {'form':form})


@log_view
@examiner_required
def create_question_category(request):

    # Create a new question category based on the form submission.
    if request.method == "POST":
        name = request.POST.get("category_name")
        description = request.POST.get("category_description")
        visible_to_others = request.POST.get("visible_to_others") == "on"

        # Validate input for name and description
        if not name or not description:
            messages.error(request, "Please provide both name and description.")
            return render(request, "QuizCreator/create_question_category.html")

        try:
            # Check for existing category with the same name for the current user
            existing_category = Category.objects.filter(name=name, created_by=request.user).exists()
            if existing_category:
                messages.error(request, "Category with the same name already exists for this user.")
                return render(request, "QuizCreator/create_question_category.html")
            
            # Create a new category
            category = Category.objects.create(
                name=name,
                description=description,
                visible_to_others=visible_to_others,
                created_by=request.user
            )
            messages.success(request, "Category created successfully.")
            
        except Exception as e:
            # Display an error message if category creation fails
            messages.error(request, f"Failed to create category: {e}")
            return render(request, "QuizCreator/create_question_category.html")
    return render(request, "QuizCreator/create_question_category.html")



@log_view
@examiner_required
def view_questions(request):
    # Display and filter questions for viewing.
    current_user = request.user

    # Retrieve categories accessible to the current user
    categories = Category.objects.filter(
        Q(created_by=current_user) | Q(visible_to_others=True)
    )

    if request.method == "POST":
        # Handle form submission to filter questions based on selected categories
        selected_categories = request.POST.getlist('category-select')  
        selected_categories = [int(cat_id) for cat_id in selected_categories]

        if selected_categories :
            # Filter questions based on selected categories
            filtered_questions = Question.objects.filter(categoryquestionmap__category_id__in=selected_categories).select_related('type').prefetch_related(
                Prefetch('options', queryset=QuestionOption.objects.all()),
                Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
            ).distinct()
        else:
            filtered_questions = Question.objects.select_related('type').prefetch_related(
                Prefetch('options', queryset=QuestionOption.objects.all()),
                Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
            ).all()

        return render(request, 'QuizCreator/questions.html', {'questions': filtered_questions,  'categories': categories, 'selected_categories': selected_categories})  
    
    # Display all questions and categories for viewing
    questions = Question.objects.select_related('type').prefetch_related(
        Prefetch('options', queryset=QuestionOption.objects.all()),
        Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
    ).all()
    return render(request, 'QuizCreator/questions.html', {'questions': questions, 'categories': categories})

 
@log_view
@examiner_required
def edit_questions(request):
    # Allow to edit the question
    questions = Question.objects.all()
    current_user = request.user
    question_types = QuestionType.objects.all()
    categories = Category.objects.filter(
        Q(created_by=current_user) | Q(visible_to_others=True)
    )   


@log_view
@examiner_required
def edit_questions(request):
    # Display and filter questions for editing.
    current_user = request.user

    # Retrieve categories accessible to the current user
    categories = Category.objects.filter(
        Q(created_by=current_user) | Q(visible_to_others=True)
    )
    if request.method == "POST":
        # Handle form submission to filter questions based on selected categories
        selected_categories = request.POST.getlist('category-select') 
        selected_categories = [int(cat_id) for cat_id in selected_categories]
        if selected_categories :
            # Filter questions based on selected categories
            filtered_questions = Question.objects.filter(categoryquestionmap__category_id__in=selected_categories).select_related('type').prefetch_related(
                Prefetch('options', queryset=QuestionOption.objects.all()),
                Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
            ).distinct()
        else:
            # Display all questions if no category is selected
            filtered_questions = Question.objects.select_related('type').prefetch_related(
                Prefetch('options', queryset=QuestionOption.objects.all()),
                Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
            ).all()
        return render(request, 'QuizCreator/edit_questions.html', {'questions': filtered_questions,  'categories': categories, 'selected_categories': selected_categories})   
     
    # Display all questions and categories for editing
    questions = Question.objects.select_related('type').prefetch_related(
        Prefetch('options', queryset=QuestionOption.objects.all()),
        Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
    ).all()
    return render(request, 'QuizCreator/edit_questions.html', {'questions': questions, 'categories': categories})

@log_view
@examiner_required
def delete_question(request, question_id):
    # Delete a question based on the provided question ID.
    try:
        # Attempt to retrieve and delete the question
        question = get_object_or_404(Question, pk=question_id)
        quizzes = QuizQuestion.objects.filter(question=question)
        for quiz_question in quizzes:
            user_quiz_status = UserQuizStatus.objects.filter(quiz=quiz_question.quiz)
            if user_quiz_status and quiz_question.quiz.visible:
                messages.error(request, "Question Can Not Be Deleted As It Is Present In The Current Active Posted Quiz")
                return redirect("edit_questions")
        question.delete()
    except Question.DoesNotExist as e:
        messages.error(f"ERROR :  Could Not Delete Question {str(e)}")

    # Display a success message and redirect to the 'edit_questions' page
    messages.success(request, "Question Has Been Deleted Successfully")
    return redirect("edit_questions")

@log_view
@examiner_required
def map_question_category(request):
    #Map selected questions to one or more categories.
    if request.method == 'POST':
        # Retrieve selected questions and categories from form submission
        checkbox_values = request.POST.getlist('checkbox-group')  
        selected_categories = request.POST.getlist('category-select') 
        # Map each selected question to the chosen categories 
        for checkbox_value in checkbox_values:
            question_id = int(checkbox_value)
            for category_id in selected_categories:
                category = Category.objects.filter(id=category_id).first()
                question = Question.objects.filter(id=question_id).first()
                CategoryQuestionMap.objects.update_or_create(
                    question=question,
                    category=category
                )
        messages.success(request, "Mapping Added Succesfully")
    # Retrieve questions and categories for display
    questions = Question.objects.select_related('type').prefetch_related(
        Prefetch('options', queryset=QuestionOption.objects.all()),
        Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
    ).all()    
    current_user = request.user
    categories = Category.objects.filter(
        Q(created_by=current_user) | Q(visible_to_others=True)
    )
    return render(request, "QuizCreator/question_category_map.html", {'questions': questions, 'categories': categories})


@log_view
@examiner_required
def create_quiz(request):

    # Create quiz option with Examineer validation
    if request.method == 'POST':
        quiz_name = request.POST.get('quiz-name')
        quiz_description = request.POST.get('quiz-description')
        quiz_duration = request.POST.get('quiz-duration')
        quiz_total_questions = request.POST.get('quiz-total-questions')
        quiz_passcode = request.POST.get('passcode')
        new_quiz = Quiz(
            name=quiz_name,
            description=quiz_description,
            duration=quiz_duration,
            total_questions=quiz_total_questions,
            passcode=quiz_passcode,
            created_by = request.user,
            modified_by = request.user
        )
        new_quiz.save()
        messages.success(request, 'Quiz created successfully!')
    return render(request, "QuizCreator/create_quiz.html")


@log_view
@examiner_required
def edit_quiz(request):

    #Display quizzes for editing and handle the selection of a quiz for further modification.
    quizzes = Quiz.objects.all()
    selected_quiz = None
    if request.method == 'POST':
        selected_quiz_id = request.POST.get('select-quiz')
        selected_quiz_id = int(selected_quiz_id)
        selected_quiz = Quiz.objects.get(id=selected_quiz_id)
    return render(request, 'QuizCreator/edit_quiz.html', {'quizzes': quizzes, 'selected_quiz': selected_quiz})


@log_view
@examiner_required
def update_quiz(request):
    #Update details of a quiz based on form submission.
    if request.method == 'POST':
        # Retrieve quiz details from the form submission
        quiz_id = request.POST.get('quiz-id')
        selected_quiz = Quiz.objects.get(pk=quiz_id)
        selected_quiz.name = request.POST.get('quiz-name')
        selected_quiz.description = request.POST.get('quiz-description')
        selected_quiz.duration = request.POST.get('quiz-duration')
        selected_quiz.total_questions = request.POST.get('quiz-total-questions')
        selected_quiz.visible = request.POST.get('quiz-visible') == 'on'
        selected_quiz.passcode = request.POST.get('quiz-passcode')
        selected_quiz.modified_by = request.user
        selected_quiz.save()

        # Provide a success message and redirect to the 'edit_quiz' page
        messages.success(request, 'Quiz details updated successfully.')
        return redirect('edit_quiz')  
    
    # Display all quizzes for editing
    quizzes = Quiz.objects.all()
    return render(request, 'QuizCreator/edit_quiz.html', {'quizzes': quizzes})

@log_view
@examiner_required
def delete_quiz(request):
    return redirect("edit_quiz")

@log_view
@examiner_required
def post_quiz(request):
    # Handle the posting of a quiz to selected groups or participants.
    quizzes = Quiz.objects.filter(created_by=request.user)
    groups = Group.objects.filter(created_by=request.user)
    context = {
        'quizzes': quizzes, 
        'groups': groups
    }
    if request.method == "POST":
        form_type = request.POST.get('form-type')

        # Handling form submission for selecting a group
        if form_type == 'select-group':
            group_id = request.POST.get('select-group')
            group = Group.objects.get(id=int(group_id))
            group_examinees = [mapping.user for mapping in GroupExamineeMapping.objects.filter(group=group)]
            context['group_examinees'] = group_examinees
            context['selected_group'] = group

        # Handling form submission for posting a quiz
        else:
            group_id = request.POST.get('group_id')
            group = Group.objects.get(id=int(group_id))
            quiz_id = request.POST.get('select-quiz')
            quiz = Quiz.objects.get(id=int(quiz_id))
            quiz_questions = QuizQuestion.objects.filter(quiz=quiz)

            group_examinees = [mapping.user for mapping in GroupExamineeMapping.objects.filter(group=group)]
            context['group_examinees'] = group_examinees

            # Check if the quiz has enough questions
            if len(quiz_questions) < quiz.total_questions:
                messages.error(request, f"Please Add Questions To The Quiz. Required Total Questions in The Quiz : {quiz.total_questions}")
                return redirect('post_quiz')
            quiz_participants = request.POST.getlist('checkbox-group')

            # Update participants' quiz status
            for user_id in quiz_participants:
                user = User.objects.get(id=user_id)
                try:

                    user_quiz_status = UserQuizStatus.objects.get(user=user, quiz=quiz)
                except UserQuizStatus.DoesNotExist:
                    user_quiz_status = None
                if not user_quiz_status:
                    # Create and save a new UserQuizStatus object
                    notification_obj = notification.Notification()
                    new_user_quiz_status = UserQuizStatus(user=user,quiz=quiz)
                    new_user_quiz_status.attach_observer(notification_obj)
                    new_user_quiz_status.save()
                    new_user_quiz_status.detach_observer(notification_obj)   

            # Update quiz visibility and context               
            quiz.visible = True
            quiz.save()
            context['selected_quiz'] = quiz
            context['selected_group'] = group
            messages.success(request, "Quiz Posted To Selected Participants")
    return render(request, 'QuizCreator/post_quiz.html', context)

@log_view
@examiner_required
def add_quiz_questions(request):
    # Initialize variables
    total_questions_added = 0
    quiz_questions = []
    quizzes = Quiz.objects.filter(created_by=request.user)
    remaining_questions = None
    selected_quiz = None
    current_user = request.user
    categories = Category.objects.filter(
        Q(created_by=current_user) | Q(visible_to_others=True)
    )
    questions = Question.objects.select_related('type').prefetch_related(
        Prefetch('options', queryset=QuestionOption.objects.all()),
        Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
    ).all()

    # Handle form submissions
    if request.method == "POST":
        form_type = request.POST.get("form_type")

        # Add questions to the selected quiz
        if form_type == "add_quiz_questions":
            quiz_id = request.POST.get("selected_quiz_id")
            selected_questions = request.POST.getlist('checkbox-group')
            try:
                # Attempt to retrieve the quiz and add selected questions
                quiz = get_object_or_404(Quiz, pk=quiz_id)
                for question_id in selected_questions:
                    existing_mapping = QuizQuestion.objects.filter(quiz=quiz, question_id=question_id).exists()
                    if not existing_mapping:
                        question = get_object_or_404(Question, pk=question_id)
                        quiz_question = QuizQuestion.objects.create(quiz=quiz, question=question)
                        quiz_question.save()
                messages.success(request, "Questions Added Successfully")

                # Update variables for displaying information on the page
                quiz_question_objects = QuizQuestion.objects.filter(quiz = quiz)
                total_questions_added = quiz_question_objects.count()
                remaining_questions = quiz.total_questions - total_questions_added
                quiz_questions = Question.objects.select_related('type').prefetch_related(
                    Prefetch('options', queryset=QuestionOption.objects.all()),
                    Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
                ).filter(quizquestion__in=quiz_question_objects).all()                
            except Quiz.DoesNotExist:
                messages.error(request, "ERROR : Invalid Quiz Id")
            selected_quiz = get_object_or_404(Quiz, pk=quiz_id)

        # Select questions for a quiz
        elif form_type == "select_quiz":
            quiz_id = request.POST.get("select-quiz")
            selected_quiz_id = quiz_id
            try:
                # Attempt to retrieve the quiz and associated questions
                quiz = get_object_or_404(Quiz, pk=quiz_id)
                quiz_question_objects = QuizQuestion.objects.filter(quiz = quiz)
                total_questions_added = quiz_question_objects.count()
                remaining_questions = quiz.total_questions - total_questions_added
                quiz_questions = Question.objects.select_related('type').prefetch_related(
                    Prefetch('options', queryset=QuestionOption.objects.all()),
                    Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
                ).filter(quizquestion__in=quiz_question_objects).all()               
            except Quiz.DoesNotExist:
                messages.error(request, "ERROR : Invalid Quiz Id")
            selected_quiz = get_object_or_404(Quiz, pk=selected_quiz_id)

        # Search questions by category
        elif form_type == "question_search_by_category":
            quiz_id = request.POST.get("selected_quiz_id")
            selected_quiz_id = quiz_id
            selected_categories = request.POST.getlist('category-select')  # Retrieve selected categories
            selected_categories = [int(cat_id) for cat_id in selected_categories]
            if selected_categories :
                # Filter questions by selected categories
                questions = Question.objects.filter(categoryquestionmap__category_id__in=selected_categories).select_related('type').prefetch_related(
                    Prefetch('options', queryset=QuestionOption.objects.all()),
                    Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
                ).distinct()
            else:
                questions = Question.objects.select_related('type').prefetch_related(
                    Prefetch('options', queryset=QuestionOption.objects.all()),
                    Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
                ).all()
            try:
                # Attempt to retrieve the quiz and update displayed information
                quiz = get_object_or_404(Quiz, pk=quiz_id)
                quiz_question_objects = QuizQuestion.objects.filter(quiz = quiz)
                total_questions_added = quiz_question_objects.count()
                remaining_questions = quiz.total_questions - total_questions_added
                quiz_questions = Question.objects.select_related('type').prefetch_related(
                    Prefetch('options', queryset=QuestionOption.objects.all()),
                    Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
                ).filter(quizquestion__in=quiz_question_objects).all()
            except Quiz.DoesNotExist:
                messages.error(request, "ERROR : Invalid Quiz Id")
            selected_quiz = get_object_or_404(Quiz, pk=selected_quiz_id)

    # Collect question IDs for display
    question_id_list = []
    for question in quiz_questions:
        question_id_list.append(question.id)

    # Ensure remaining_questions is non-negative
    if remaining_questions:
        if remaining_questions < 0:
            remaining_questions = 0

    # Render the page with collected information
    return render(request, 'QuizCreator/modify_quiz_questions.html', {'quizzes': quizzes, 'categories' : categories, 'questions' : questions, 'quiz_questions': quiz_questions, 'total_questions_added' : total_questions_added, 'selected_quiz' : selected_quiz, 'question_id_list' : question_id_list, 'remaining_questions' : remaining_questions})

@log_view
@examiner_required
def make_quiz_visible(request):
    if request.method == "POST":
        quiz_id = request.POST.get('quiz_id')
        try:
            # Attempt to retrieve the quiz and update its visibility status
            quiz = get_object_or_404(Quiz, pk=quiz_id)
            quiz.visible = True
            quiz.save()
            messages.success(request, "Quiz Has Been Posted Successfully")
        except Quiz.DoesNotExist:
            messages.error(request, "Error: Could Not Made Quiz Available. Try Again")
    # Redirect to the page for adding quiz questions
    return redirect('add_quiz_questions')

@log_view
@examiner_required
def delete_quiz_questions(request):
    if request.method == "POST":
        # Retrieve the list of question IDs and the associated quiz ID from the form submission
        question_ids = request.POST.getlist("delete-checkbox-group")
        quiz_id = request.POST.get("selected_quiz_id")
        try:
            # Attempt to retrieve the quiz and delete the selected questions
            quiz = Quiz.objects.get(pk=quiz_id)
            for question_id in question_ids:
                question = Question.objects.get(pk=question_id)
                QuizQuestion.objects.filter(quiz=quiz, question=question).delete()
            messages.success(request, "Questions Deleted")
        except (Quiz.DoesNotExist, Question.DoesNotExist) as e:
            messages.error(request, "ERROR : Could Not Delete ")
        # Redirect to the page for adding quiz questions 
        return redirect("add_quiz_questions")
    

@log_view
@examiner_required
def preview_quiz(request):
    quizzes = Quiz.objects.filter(created_by=request.user)
    if request.method == "POST":
        quiz_id = request.POST.get("select-quiz")
        try:
            # Retrieve the selected quiz and associated questions
            quiz = get_object_or_404(Quiz, pk=quiz_id)
            quiz_question_objects = QuizQuestion.objects.filter(quiz = quiz)
            quiz_questions = Question.objects.select_related('type').prefetch_related(
                Prefetch('options', queryset=QuestionOption.objects.all()),
                Prefetch('categoryquestionmap_set', queryset=CategoryQuestionMap.objects.select_related('category'))
            ).filter(quizquestion__in=quiz_question_objects).all()
            # Shuffle the questions and their options for preview
            shuffled_questions = list(quiz_questions)
            random.shuffle(shuffled_questions)
            for question in shuffled_questions:
                question_options = list(question.options.all())
                random.shuffle(question_options)
                question.options.set(question_options)
            shuffled_questions = shuffled_questions[:quiz.total_questions]            
        except Quiz.DoesNotExist:
            messages.error(request, "ERROR : Invalid Quiz Id")
        return render(request, "QuizCreator/show_preview.html", {'quiz' : quiz,'quiz_questions' : shuffled_questions})
    # Render the page for selecting a quiz
    return render(request, "QuizCreator/preview_quiz.html", {'quizzes' : quizzes})


@log_view
@examiner_required
def add_examinee_to_group(request):
    # Retrieve the groups created by the current user
    group_id = None
    groups = Group.objects.filter(created_by=request.user)   
    # Retrieve all users marked as examinees
    examinees = User.objects.filter(is_examinee=True)
    if request.method == 'POST':
        form_type = request.POST.get('form-type')        
        # Handle form submission for selecting a group
        if form_type == 'select-group':
            group_id = int(request.POST.get('select-group'))
            group = Group.objects.get(id=group_id)          
            # Retrieve examinees already mapped to the selected group
            group_examinees = GroupExamineeMapping.objects.filter(group=group)
            examinee_ids = []
            for group_examinee in group_examinees:
                examinee_ids.append(group_examinee.user.id)
            # Render the page with relevant information
            return render(request, 'QuizCreator/add_examinee_to_group.html', {'group_id': group_id, 'examinees': examinees, 'groups': groups, 'examinee_ids': examinee_ids})
        # Handle form submission for adding examinees to a group
        elif form_type == 'add-examinees':
            group_id = request.POST.get('selected-group')
            selected_examinees = request.POST.getlist('checkbox-group')
            group = Group.objects.get(id=group_id)         
            # Remove existing mappings that are not in the selected list
            existing_mapping = GroupExamineeMapping.objects.filter(group=group)
            for existing in existing_mapping:
                id = str(existing.user.id)
                if id not in selected_examinees:
                    existing.delete()        
            # Add new mappings for selected examinees
            for examinee_id in selected_examinees:
                examinee = User.objects.get(id=examinee_id)
                GroupExamineeMapping.objects.get_or_create(group=group, user=examinee)  
            messages.success(request, 'Group updated successfully!') 
            # Render the page with updated information
            return render(request, 'QuizCreator/add_examinee_to_group.html', {'examinees': examinees, 'groups': groups})
    # Render the initial page with group and examinee information
    return render(request, 'QuizCreator/add_examinee_to_group.html', {'group_id': group_id, 'examinees': examinees, 'groups': groups})


@log_view
@examiner_required
# Creates a new group with user-provided details or displays the group creation form.
def create_group(request):
    if request.method == 'POST':
        group_name = request.POST.get('group-name')
        group_description = request.POST.get('group-description')
        new_group = Group(
            name=group_name,
            description=group_description,
            created_by = request.user
        )
        new_group.save()
        messages.success(request, 'Group created successfully!')
        return redirect('create_group')
    return render(request, 'QuizCreator/create_group.html')


@log_view
@examiner_required
# retrieves and displays examinees associated with the selected group.
def view_group(request):
    groups = Group.objects.filter(created_by=request.user)
    selected_group_id = None
    examinees = []
    if request.method == 'POST':
        selected_group_id = request.POST.get('select-group')
        selected_group = Group.objects.get(id=selected_group_id)
        mappings = GroupExamineeMapping.objects.filter(group=selected_group)
        examinees = []
        for mapping in mappings:
            examinees.append(mapping.user)
        selected_group_id = int(selected_group_id)
        return render(request, "QuizCreator/view_group.html", {"examinees" : examinees, 'groups': groups, 'group_id': selected_group_id})
    return render(request, 'QuizCreator/view_group.html', {
        'groups': groups,
        'selected_group_id': selected_group_id,
        'examinees': examinees,
    })